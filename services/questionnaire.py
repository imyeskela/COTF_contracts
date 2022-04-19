import random
import time
from io import BytesIO
from django.utils import timezone
from django.core.files.base import ContentFile

import os
from docxtpl import DocxTemplate
from docxtpl import InlineImage

from PIL import Image as PilImage


from babel.dates import format_date
from docx.shared import Mm, Inches, Pt
import qrcode
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from django.core.mail import EmailMessage
import base64
import convertapi
import docx2txt
from twilio.rest import Client

from cotf_contracts.settings import BASE_DIR, EMAIL_HOST_USER, CONVERT_API_SECRET, MEDIA_URL, AUTH_TOKEN, ACCOUNT_SID
from services.main_logic import get_contract, get_codes_of_obj, get_num_attempts, get_code_obj
from services.num_to_text import num2text
from main.models import AuthenticationCode

import mammoth


def get_data_from_forms(self, request, contract_number):
    if request.method == 'POST':
        contract = get_contract(self)
        contract_pk = contract.pk
        last_name = request.POST['last_name']
        name = request.POST['name']
        sur_name = request.POST['sur_name']
        full_name = last_name.title() + ' ' + name.title() + ' ' + sur_name.title()
        date_created = contract.date_created
        generated_date = format_date(date_created, 'd MMMM yyyy', locale='ru')
        phone = request.POST['phone']
        passport = request.POST['series_passport'] + request.POST['num_passport']
        email = request.POST['email']
        company = contract.contract_template.company
        if sur_name:
            short_name = last_name + ' ' + name[0] + '.' + sur_name[0] + '.'
        else:
            short_name = last_name + ' ' + name[0] + '.'
        sum = contract.amount
        text_sum = num2text(sum) + ' рублей'
        id_contract = str(contract.number)
        type = str(contract.contract_template.type)
        contract_template_path = contract.contract_template.template_of_contract.path
        return {'last_name': last_name, 'name': name, 'sur_name': sur_name, 'full_name': full_name,
                'generated_date': generated_date, 'phone': phone, 'passport': passport, 'email': email,
                'short_name': short_name, 'id_contract': id_contract, 'type': type, 'sum_c': sum,
                'text_sum': text_sum, 'contract_template_path': contract_template_path, 'company': company,
                'date_created': date_created, 'contract_pk': contract_pk}


def basic_vars(self, request, contract_number):
    vars = {'email': get_data_from_forms(self, request, contract_number).get('email'),
            'full_name': get_data_from_forms(self, request, contract_number).get('full_name'),
            'id': get_data_from_forms(self, request, contract_number).get('id_contract'),
            'short_name': get_data_from_forms(self, request, contract_number).get('short_name'),
            'generated_date': str(get_data_from_forms(self, request, contract_number).get('generated_date')),
            'phone': get_data_from_forms(self, request, contract_number).get('phone'),
            'passport': get_data_from_forms(self, request, contract_number).get('passport'),
            }
    return vars


def renewal_vars(self, request, contract_number):
    vars = {'email': get_data_from_forms(self, request, contract_number).get('email'),
            'full_name': get_data_from_forms(self, request, contract_number).get('full_name'),
            'id': get_data_from_forms(self, request, contract_number).get('id_contract'),
            'short_name': get_data_from_forms(self, request, contract_number).get('short_name'),
            'generated_date': str(get_data_from_forms(self, request, contract_number).get('generated_date')),
            'phone': get_data_from_forms(self, request, contract_number).get('phone'),
            'passport': get_data_from_forms(self, request, contract_number).get('passport'),
            'sum': get_data_from_forms(self, request, contract_number).get('sum_c'),
            'text_sum': get_data_from_forms(self, request, contract_number).get('text_sum')
            }
    return vars


def create_docx(self, request, contract_number):
    contract = get_contract(self)
    docx = DocxTemplate(os.path.join(BASE_DIR, contract.contract_template.template_of_contract.path))
    path_docx = BytesIO()
    if get_data_from_forms(self, request, contract_number).get('type') == 'Основной':

        docx.render(basic_vars(self, request, contract_number))

        docx.save(path_docx)
    elif get_data_from_forms(self, request, contract_number).get('type') == 'Продление':

        docx.render(renewal_vars(self, request, contract_number))
        docx.save(path_docx)
    path_docx.seek(0)
    return path_docx


def form_questionnaire(self, request, contract_number):
    docx = create_docx(self, request, contract_number)
    html = mammoth.convert_to_html(docx)
    result = html.value
    return result


def get_sign_img(request):
    img = request.POST['sign']
    img = img[23:]
    return img

def change_contract_status(self):
    contract = get_contract(self)
    contract.status = 'Подписан'
    contract.save()
    return contract

def finally_rich(self, request, contract_number):
    global purpose, sum
    contract = get_contract(self)
    company = get_data_from_forms(self, request, contract_number).get('company')
    generated_date_qr = format_date(get_data_from_forms(self, request, contract_number).get('date_created'), 'd.MM.yy', locale='ru')
    qr = qrcode.QRCode(
        version=4,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=2,
        border=1,
    )
    id_contract = str(get_data_from_forms(self, request, contract_number).get('id_contract'))
    type = get_data_from_forms(self, request, contract_number).get('type')
    if type == 'Основной':
        purpose = 'Оплата участия в серии мероприятий для бизнеса "Клуба Первых". ' \
                  'По Договору оказания услуг ФЛМ-' + id_contract + ' от ' + generated_date_qr
        sum = get_data_from_forms(self, request, contract_number).get('sum_c')
    elif type == 'Продление':
        purpose = 'Оплата участия в серии мероприятий для бизнеса "Клуба Первых" ' \
                  'Тариф "Продление" по Договору оказания услуг ПМ-' + id_contract + ' от ' + generated_date_qr
        sum = get_data_from_forms(self, request, contract_number).get('sum_c')

    new_path_docx = os.path.join(BASE_DIR, 'upload/signed_contract/' + id_contract + '.docx')
    path_pdf = os.path.join(BASE_DIR, 'upload/signed_contract/' + id_contract + '.pdf')
    last_name = get_data_from_forms(self, request, contract_number).get('last_name')
    name = get_data_from_forms(self, request, contract_number).get('name')
    sur_name = get_data_from_forms(self, request, contract_number).get('sur_name')
    email = get_data_from_forms(self, request, contract_number).get('email')
    docx = DocxTemplate(os.path.join(BASE_DIR, contract.contract_template.template_of_contract.path))

    basic_vars = {
        'email': get_data_from_forms(self, request, contract_number).get('email'),
        'full_name': get_data_from_forms(self, request, contract_number).get('full_name'),
        'id': get_data_from_forms(self, request, contract_number).get('id_contract'),
        'short_name': get_data_from_forms(self, request, contract_number).get('short_name'),
        'generated_date': str(get_data_from_forms(self, request, contract_number).get('generated_date')),
        'phone': get_data_from_forms(self, request, contract_number).get('phone'),
        'passport': get_data_from_forms(self, request, contract_number).get('passport'),
    }


    renewal_vars = {
        'email': get_data_from_forms(self, request, contract_number).get('email'),
        'full_name': get_data_from_forms(self, request, contract_number).get('full_name'),
        'id': get_data_from_forms(self, request, contract_number).get('id_contract'),
        'short_name': get_data_from_forms(self, request, contract_number).get('short_name'),
        'generated_date': str(get_data_from_forms(self, request, contract_number).get('generated_date')),
        'phone': get_data_from_forms(self, request, contract_number).get('phone'),
        'passport': get_data_from_forms(self, request, contract_number).get('passport'),
        'sum': get_data_from_forms(self, request, contract_number).get('sum_c'),
        'text_sum': get_data_from_forms(self, request, contract_number).get('text_sum')
    }

    if company == 'ООО "КМС"':

        qr_kms = qr
        qr_kms.add_data(
            'ST00012|Name=ООО "КМС"|'
            'PersonalAcc=40702810838000108799|'
            'BankName=ПАО СБЕРБАНК|'
            'BIC=044525225|'
            'CorrespAcc=30101810400000000225|'
            'Purpose=' + str(purpose) + '|'
            'Sum=' + str(sum) + '|'
            'PayeeINN=7725731508|'
            'KPP=772501001|'
            'LastName=' + last_name + '|'
            'FirstName=' + name + '|'
            'MiddleName=' + sur_name + ''
        )

        qr_kms.make(fit=True)

        img = qr_kms.make_image(fill_color="black", back_color="white")
        bytes_io = BytesIO()
        img.save(bytes_io, format='jpeg')
        img_base = base64.b64encode(bytes_io.getvalue()).decode()

        workbook = load_workbook(os.path.join(BASE_DIR, 'Шаблон Счета на оплату КМС.xlsx'))
        sheet = workbook.active
        sheet['C8'] = purpose
        sheet['C10'] = str(sum)
        sheet['C21'] = purpose
        sheet['C23'] = str(sum)
        img_qr = Image(bytes_io)
        sheet.add_image(img_qr, 'B18')

        path_payment = os.path.join(BASE_DIR, 'upload/payment/счет_' + id_contract + '.xlsx')
        workbook.save(path_payment)
        sign_img_base = get_sign_img(request)
        sign_img = ContentFile(base64.b64decode(sign_img_base))
        if type == 'Основной':

            basic_vars['signature'] = InlineImage(docx, sign_img, width=Mm(30), height=Mm(20))
            docx.render(basic_vars)

            docx.save(new_path_docx)
        else:
            renewal_vars['signature'] = InlineImage(docx, sign_img, width=Mm(30), height=Mm(20))
            docx.render(renewal_vars)

            docx.save(new_path_docx)
        convertapi.api_secret = CONVERT_API_SECRET
        result = convertapi.convert('pdf', {'File': new_path_docx})

        result.file.save(path_pdf)
        os.remove(new_path_docx)
        contract.identifier = get_data_from_forms(self, request, contract_number).get('short_name')
        contract.full_name = f'{last_name} {name} {sur_name}'
        contract.email = email
        contract.series_passport = get_data_from_forms(self, request, contract_number).get('series_passport')
        contract.num_passport = get_data_from_forms(self, request, contract_number).get('num_passport')
        contract.phone = get_data_from_forms(self, request, contract_number).get('phone')
        contract.payment = path_payment
        contract.signed_contract = path_pdf
        contract.date_signed = timezone.now()
        contract.save()
        # email_send = EmailMessage(
        #     subject='Клуб Первых',
        #     body='Оплатить счёт возможно в течение 5 рабочих дней. До встречи в Клубе Первых!',
        #     from_email=EMAIL_HOST_USER,
        #     to=[str(email)],
        # )
        # email_send.attach_file(path_payment)
        # email_send.attach_file(path_pdf)
        # email_send.send()

        return img_base
    elif company == 'ООО "ДЕЛОВОЙ КЛУБ"':
        qr_dk = qr
        qr_dk.add_data(
            'ST00012|Name=ООО "ДЕЛОВОЙ КЛУБ"|'
            'PersonalAcc=40702810338000156966|'
            'BankName=ПАО СБЕРБАНК|'
            'BIC=044525225|'
            'CorrespAcc=30101810400000000225|'
            'Purpose=' + purpose + '|'
                                   'Sum=' + str(sum) + '|'
                                                       'PayeeINN=9715357887|'
                                                       'KPP=771501001|'
                                                       'LastName=' + last_name + '|'
                                                                                 'FirstName=' + name + '|'
                                                                                                       'MiddleName=' + sur_name + ''
        )

        qr_dk.make(fit=True)
        img = qr_dk.make_image(fill_color="black", back_color="white")
        bytes_io = BytesIO()
        img.save(bytes_io, format='jpeg')
        img_base = base64.b64encode(bytes_io.getvalue()).decode()
        workbook = load_workbook(os.path.join(BASE_DIR, 'Шаблон Счета на оплату ДК.xlsx'))
        sheet = workbook.active
        sheet['C8'] = purpose
        sheet['C10'] = str(sum)
        sheet['C21'] = purpose
        sheet['C23'] = str(sum)
        img_qr = Image(bytes_io)
        sheet.add_image(img_qr, 'B18')

        path_payment = os.path.join(BASE_DIR, 'upload/payment/счет_' + id_contract + '.xlsx')
        workbook.save(path_payment)
        sign_img_base = get_sign_img(request)
        sign_img = ContentFile(base64.b64decode(sign_img_base))
        if type == 'Основной':
            basic_vars['signature'] = InlineImage(docx, sign_img, width=Mm(30), height=Mm(20))
            docx.render(basic_vars)
            docx.save(new_path_docx)
        else:
            renewal_vars['signature'] = InlineImage(docx, sign_img, width=Mm(30), height=Mm(20))
            docx.render(renewal_vars)
            docx.save(new_path_docx)

        convertapi.api_secret = CONVERT_API_SECRET

        result = convertapi.convert('pdf', {'File': new_path_docx})

        result.file.save(path_pdf)
        os.remove(new_path_docx)
        contract.identifier = get_data_from_forms(self, request, contract_number).get('short_name')
        contract.full_name = f'{last_name} {name} {sur_name}'
        contract.email = email
        contract.series_passport = get_data_from_forms(self, request, contract_number).get('series_passport')
        contract.num_passport = get_data_from_forms(self, request, contract_number).get('num_passport')
        contract.phone = get_data_from_forms(self, request, contract_number).get('phone')
        contract.payment = path_payment
        contract.date_signed = timezone.now()
        contract.signed_contract = path_pdf
        contract.save()
        # email_send = EmailMessage(
        #     subject='Клуб Первых',
        #     body='Оплатить счёт возможно в течение 5 рабочих дней. До встречи в Клубе Первых!',
        #     from_email=EMAIL_HOST_USER,
        #     to=[str(email)],
        # )
        # email_send.attach_file(path_payment)
        # email_send.attach_file(path_pdf)
        # email_send.send()
        return img_base


def send_email_contract_signed(self, request, contract_number):
    full_name = get_data_from_forms(self, request, contract_number).get('full_name')
    # email = get_data_from_forms(self, request, contract_number).get('email')
    #
    # email_send = EmailMessage(
    #     subject='Клуб Первых',
    #     body=full_name + ' подписал(а) договор',
    #     from_email=EMAIL_HOST_USER,
    #     to=[str(email)],
    # )
    # email_send.send()
    return full_name

def _create_new_code():
    codes = get_codes_of_obj()
    new_code = int(''.join(random.sample('0123456789', 5)))
    if new_code in codes:
        new_code = int(''.join(random.sample('0123456789', 5)))
    return new_code


def get_num_of_attempts(self):
    attempts = len(get_num_attempts(self))
    return attempts


def create_num_attempts(self):
    attempts = get_num_of_attempts(self)
    choice_attempts = range(0, 4)
    c = get_code_obj(self).last()
    if attempts == 0:
        new_attempts = choice_attempts[0]
    elif attempts == 1:
        new_attempts = choice_attempts[1]
        c.relevance = False
        c.save()
    elif attempts == 2:
        c.relevance = False
        c.save()
        new_attempts = choice_attempts[2]
    else:
        c.relevance = False
        c.save()
        new_attempts = choice_attempts[3]

    return new_attempts


def get_actual_code(self, phone):
    try:
        return AuthenticationCode.objects.get(phone=phone, contract=get_contract(self), relevance=True)
    except:
        return None



def create_new_code_obj(self, request, contract_number):
    phone = get_data_from_forms(self, request, contract_number).get('phone')
    contract_pk = get_data_from_forms(self, request, contract_number).get('contract_pk')
    new_attempts = create_num_attempts(self)
    new_code = _create_new_code()
    new_obj = AuthenticationCode.objects.create(code=new_code,
                                                phone=phone,
                                                contract_id=contract_pk,
                                                date_generated_code=str(time.time()),
                                                confirmation=False,
                                                relevance=True,
                                                number_of_attempts=new_attempts
                                                )

    return new_obj


def send_sms(self, request, contract_number):
    phone = get_data_from_forms(self, request, contract_number).get('phone')
    # account_sid = ACCOUNT_SID
    # auth_token = AUTH_TOKEN
    # client = Client(account_sid, auth_token)
    #
    # message = client.messages.create(
    #     body=str(get_actual_code(self, phone=phone)),
    #     from_='+19378842345',
    #     to=str(phone)
    # )
    print(get_actual_code(self, phone=phone))
    return phone


def get_time_for_resend_sms(self, request, contract_number):
    actual_code = get_actual_code(self, phone=get_data_from_forms(self, request, contract_number).get('phone'))
    unix_time = float(actual_code.date_generated_code)

    attempts = actual_code.number_of_attempts

    if attempts is None:
        resend_time = float(time.time())
    elif attempts == 0:
        resend_time = unix_time + 30.0
    elif attempts == 1:
        resend_time = unix_time + 120.0
    elif attempts == 2:
        resend_time = unix_time + 600.0
    else:
        resend_time = unix_time + 3600.0

    return resend_time


def change_confirmation(self, request, contract_number):
    actual_code = get_actual_code(self, phone=get_data_from_forms(self, request, contract_number).get('phone'))
    actual_code.confirmation = True
    actual_code.save()
    return actual_code
