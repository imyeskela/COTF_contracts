from io import BytesIO
from django.utils import timezone
from django.core.files.base import ContentFile

from django.shortcuts import render, redirect
import os
from docxtpl import DocxTemplate
from docxtpl import InlineImage
from babel.dates import format_date
from PIL import Image
from docx.shared import Mm, Inches, Pt
import qrcode
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

from django.core.mail import EmailMessage
import base64
import convertapi
import docx2txt
from cotf_contracts.settings import BASE_DIR, EMAIL_HOST_USER, CONVERT_API_SECRET
from services.main_logic import generator_num_contract
from services.num_to_text import num2text
from services.main_logic import get_contract
from django.core.paginator import Paginator


class ContractTemplateListAndCreateContractMixin:
    """Миксин для отображения всех шаблонов контрактов"""

    queryset = None
    template_name = None
    form_contract = None
    form_contract_template = None

    def get(self, request):
        paginator = Paginator(self.queryset, 5)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        return render(request, self.template_name, {'contract_template_list': page_obj,
                                                    'form_contract': self.form_contract,
                                                    'form_contract_template': self.form_contract_template
                                                    })

    def post(self, request):
        form_contract = self.form_contract(request.POST)
        form_contract_template = self.form_contract_template(request.POST, request.FILES)
        if 'form_contract' in request.POST:
            if form_contract.is_valid():
                contact_template_id = int(request.POST.get('contract_template'))
                contract_template = self.queryset.get(id=contact_template_id)
                form_contract = form_contract.save(commit=False)
                form_contract.amount = int(request.POST.get('amount_bitch'))
                form_contract.contract_template = contract_template
                form_contract.number = generator_num_contract()
                print('fc')
                form_contract.save()
                return redirect('contract_list')
            else:
                return render(request, self.template_name, {'form_contract': form_contract, 'form_contract_template': form_contract_template})
        if 'form_contract_template' in request.POST:
            if form_contract_template.is_valid():
                form_contract_template = form_contract_template.save(commit=False)
                form_contract_template.save()
                return redirect('contract_template_list')
            else:
                return render(request, self.template_name,
                              {'form_contract_template': form_contract_template})

        return render(request, self.template_name, {'form_contract': form_contract, 'form_contract_template': form_contract_template})





class ContractListMixin:
    """Миксин для отображения всех контрактов"""

    queryset = None
    template_name = None

    def get(self, request):
        paginator = Paginator(self.queryset, 5)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        template_name = self.template_name
        return render(request, template_name, {'contract_list': page_obj})


def get_data_from_forms(self, request, contract_number):
    if request.method == 'POST':
        contract = get_contract(self)
        last_name = request.POST['last_name']
        name = request.POST['name']
        sur_name = request.POST['sur_name']
        full_name = last_name.title() + ' ' + name.title() + ' ' + sur_name.title()
        date_created = contract.date_created
        generated_date = format_date(date_created, 'd MMMM yyyy', locale='ru')
        phone = request.POST['phone']
        passport = request.POST['passport']
        email = request.POST['email']
        company = contract.contract_template.company
        if sur_name:
            short_name = last_name + ' ' + name[0] + '.' + sur_name[0] + '.'
        else:
            short_name = last_name + ' ' + name[0] + '.'
        sum = contract.amount
        text_sum = num2text(sum) + ' рублей'
        id_contract = str(contract.number)
        type = str(contract.contract_template.type)
        contract_template_path = contract.contract_template.template_of_contract.path
        return {'last_name': last_name, 'name': name, 'sur_name': sur_name, 'full_name': full_name,
                'generated_date': generated_date, 'phone': phone, 'passport': passport, 'email': email,
                'short_name': short_name, 'id_contract': id_contract, 'type': type, 'sum_c': sum,
                'text_sum': text_sum, 'contract_template_path': contract_template_path, 'company': company,
                'date_created': date_created}


def basic_vars(self, request, contract_number):
    vars = {'email': get_data_from_forms(self, request, contract_number).get('email'),
                  'full_name': get_data_from_forms(self, request, contract_number).get('full_name'),
                  'id': get_data_from_forms(self, request, contract_number).get('id_contract'),
                  'short_name': get_data_from_forms(self, request, contract_number).get('short_name'),
                  'generated_date': str(get_data_from_forms(self, request, contract_number).get('generated_date')),
                  'phone': get_data_from_forms(self, request, contract_number).get('phone'),
                  'passport': get_data_from_forms(self, request, contract_number).get('passport'),
                  }
    return vars


def renewal_vars(self, request, contract_number):
    vars = {'email': get_data_from_forms(self, request, contract_number).get('email'),
                    'full_name': get_data_from_forms(self, request, contract_number).get('full_name'),
                    'id': get_data_from_forms(self, request, contract_number).get('id_contract'),
                    'short_name': get_data_from_forms(self, request, contract_number).get('short_name'),
                    'generated_date': str(get_data_from_forms(self, request, contract_number).get('generated_date')),
                    'phone': get_data_from_forms(self, request, contract_number).get('phone'),
                    'passport': get_data_from_forms(self, request, contract_number).get('passport'),
                    'sum': get_data_from_forms(self, request, contract_number).get('sum_c'),
                    'text_sum': get_data_from_forms(self, request, contract_number).get('text_sum')
                    }
    return vars


def create_docx(self, request, contract_number):
    contract = get_contract(self)
    docx = DocxTemplate(os.path.join(BASE_DIR, contract.contract_template.template_of_contract.path))
    path_docx = BytesIO()
    if get_data_from_forms(self, request, contract_number).get('type') == 'Основной':

        docx.render(basic_vars(self, request, contract_number))

        docx.save(path_docx)
    elif get_data_from_forms(self, request, contract_number).get('type') == 'Продление':

        docx.render(renewal_vars(self, request, contract_number))
        docx.save(path_docx)
    path_docx.seek(0)
    return path_docx


def form_questionnaire(self, request, contract_number):
    docx = create_docx(self, request, contract_number)
    text = docx2txt.process(docx)
    return text


def get_sign_img(self, request, contract_number):
    if request.method == 'POST':
        img = request.POST['sign']
        img = img[22:]
        return img


def finally_rich(self, request, contract_number):
    global purpose, sum
    contract = get_contract(self)
    company = get_data_from_forms(self, request, contract_number).get('company')
    generated_date_qr = format_date(get_data_from_forms(self, request, contract_number).get('date_created'), 'd.MM.yy', locale='ru')
    qr = qrcode.QRCode(
        version=4,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=2,
        border=1,
    )
    id_contract = str(get_data_from_forms(self, request, contract_number).get('id_contract'))
    type = get_data_from_forms(self, request, contract_number).get('type')
    if type == 'Основной':
        purpose = 'Оплата участия в серии мероприятий для бизнеса "Клуба Первых". ' \
                  'По Договору оказания услуг ФЛМ-' + id_contract + ' от ' + generated_date_qr
        sum = get_data_from_forms(self, request, contract_number).get('sum_c')
    elif type == 'Продление':
        purpose = 'Оплата участия в серии мероприятий для бизнеса "Клуба Первых" ' \
                  'Тариф "Продление" по Договору оказания услуг ПМ-' + id_contract + ' от ' + generated_date_qr
        sum = get_data_from_forms(self, request, contract_number).get('sum_c')

    new_path_docx = os.path.join(BASE_DIR, 'upload\\signed_contract\\' + id_contract + '.docx')
    path_pdf = os.path.join(BASE_DIR, 'upload\\signed_contract\\' + id_contract + '.pdf')
    last_name = get_data_from_forms(self, request, contract_number).get('last_name')
    name = get_data_from_forms(self, request, contract_number).get('name')
    sur_name = get_data_from_forms(self, request, contract_number).get('sur_name')
    email = get_data_from_forms(self, request, contract_number).get('email')
    docx = DocxTemplate(os.path.join(BASE_DIR, contract.contract_template.template_of_contract.path))

    basic_vars = {
        'email': get_data_from_forms(self, request, contract_number).get('email'),
        'full_name': get_data_from_forms(self, request, contract_number).get('full_name'),
        'id': get_data_from_forms(self, request, contract_number).get('id_contract'),
        'short_name': get_data_from_forms(self, request, contract_number).get('short_name'),
        'generated_date': str(get_data_from_forms(self, request, contract_number).get('generated_date')),
        'phone': get_data_from_forms(self, request, contract_number).get('phone'),
        'passport': get_data_from_forms(self, request, contract_number).get('passport'),
    }


    renewal_vars = {
        'email': get_data_from_forms(self, request, contract_number).get('email'),
        'full_name': get_data_from_forms(self, request, contract_number).get('full_name'),
        'id': get_data_from_forms(self, request, contract_number).get('id_contract'),
        'short_name': get_data_from_forms(self, request, contract_number).get('short_name'),
        'generated_date': str(get_data_from_forms(self, request, contract_number).get('generated_date')),
        'phone': get_data_from_forms(self, request, contract_number).get('phone'),
        'passport': get_data_from_forms(self, request, contract_number).get('passport'),
        'sum': get_data_from_forms(self, request, contract_number).get('sum_c'),
        'text_sum': get_data_from_forms(self, request, contract_number).get('text_sum')
    }

    if company == 'ООО "КМС"':

        qr_kms = qr
        qr_kms.add_data(
            'ST00012|Name=ООО "КМС"|'
            'PersonalAcc=40702810838000108799|'
            'BankName=ПАО СБЕРБАНК|'
            'BIC=044525225|'
            'CorrespAcc=30101810400000000225|'
            'Purpose=' + str(purpose) + '|'
            'Sum=' + str(sum) + '|'
            'PayeeINN=7725731508|'
            'KPP=772501001|'
            'LastName=' + last_name + '|'
            'FirstName=' + name + '|'
            'MiddleName=' + sur_name + ''
        )

        qr_kms.make(fit=True)

        img = qr_kms.make_image(fill_color="black", back_color="white")
        bytes_io = BytesIO()
        img.save(bytes_io, format='png')
        img_base = base64.b64encode(bytes_io.getvalue()).decode()

        workbook = load_workbook(os.path.join(BASE_DIR, 'Шаблон Счета на оплату КМС.xlsx'))
        sheet = workbook.active
        sheet['C8'] = purpose
        sheet['C10'] = str(sum)
        sheet['C21'] = purpose
        sheet['C23'] = str(sum)
        img_qr = Image(bytes_io)
        sheet.add_image(img_qr, 'B18')

        path_payment = os.path.join(BASE_DIR, 'upload\\payment\\счет_' + id_contract + '.xlsx')
        workbook.save(path_payment)
        sign_img_base = get_sign_img(self, request, contract_number)
        sign_img = ContentFile(base64.b64decode(sign_img_base))
        if type == 'Основной':

            basic_vars['signature'] = InlineImage(docx, sign_img, width=Mm(30), height=Mm(20))
            docx.render(basic_vars)

            docx.save(new_path_docx)
        else:
            renewal_vars['signature'] = InlineImage(docx, sign_img, width=Mm(30), height=Mm(20))
            docx.render(renewal_vars)

            docx.save(new_path_docx)
        convertapi.api_secret = CONVERT_API_SECRET
        result = convertapi.convert('pdf', {'File': new_path_docx})

        result.file.save(path_pdf)
        os.remove(new_path_docx)
        contract.identifier = get_data_from_forms(self, request, contract_number).get('short_name')
        contract.email = email
        contract.passport = get_data_from_forms(self, request, contract_number).get('passport')
        contract.phone = get_data_from_forms(self, request, contract_number).get('phone')
        contract.payment = path_payment
        contract.signed_contract = path_pdf
        contract.date_signed = timezone.now()
        contract.save()
        email_send = EmailMessage(
            subject='Клуб Первых',
            body='Оплатить счёт возможно в течение 5 рабочих дней. До встречи в Клубе Первых!',
            from_email=EMAIL_HOST_USER,
            to=[str(email)],
        )
        email_send.attach_file(path_payment)
        email_send.attach_file(path_pdf)
        email_send.send()
        return img_base
    elif company == 'ООО "ДЕЛОВОЙ КЛУБ"':
        qr_dk = qr
        qr_dk.add_data(
            'ST00012|Name=ООО "ДЕЛОВОЙ КЛУБ"|'
            'PersonalAcc=40702810338000156966|'
            'BankName=ПАО СБЕРБАНК|'
            'BIC=044525225|'
            'CorrespAcc=30101810400000000225|'
            'Purpose=' + purpose + '|'
                                   'Sum=' + str(sum) + '|'
                                                       'PayeeINN=9715357887|'
                                                       'KPP=771501001|'
                                                       'LastName=' + last_name + '|'
                                                                                 'FirstName=' + name + '|'
                                                                                                       'MiddleName=' + sur_name + ''
        )

        qr_dk.make(fit=True)
        img = qr_dk.make_image(fill_color="black", back_color="white")
        bytes_io = BytesIO()
        img.save(bytes_io, format='png')
        img_base = base64.b64encode(bytes_io.getvalue()).decode()
        workbook = load_workbook(os.path.join(BASE_DIR, 'Шаблон Счета на оплату ДК.xlsx'))
        sheet = workbook.active
        sheet['C8'] = purpose
        sheet['C10'] = str(sum)
        sheet['C21'] = purpose
        sheet['C23'] = str(sum)
        img_qr = Image(bytes_io)
        sheet.add_image(img_qr, 'B18')

        path_payment = os.path.join(BASE_DIR, 'upload\\payment\\счет_' + id_contract + '.xlsx')
        workbook.save(path_payment)
        sign_img_base = get_sign_img(self, request, contract_number)
        sign_img = ContentFile(base64.b64decode(sign_img_base))
        if type == 'Основной':
            basic_vars['signature'] = InlineImage(docx, sign_img, width=Mm(30), height=Mm(20))
            docx.render(basic_vars)
            docx.save(new_path_docx)
        else:
            renewal_vars['signature'] = InlineImage(docx, sign_img, width=Mm(30), height=Mm(20))
            docx.render(renewal_vars)
            docx.save(new_path_docx)

        convertapi.api_secret = CONVERT_API_SECRET

        result = convertapi.convert('pdf', {'File': new_path_docx})

        result.file.save(path_pdf)
        os.remove(new_path_docx)
        contract.identifier = get_data_from_forms(self, request, contract_number).get('short_name')
        contract.email = email
        contract.passport = get_data_from_forms(self, request, contract_number).get('passport')
        contract.phone = get_data_from_forms(self, request, contract_number).get('phone')
        contract.payment = path_payment
        contract.date_signed = timezone.now()
        contract.signed_contract = path_pdf
        contract.save()
        email_send = EmailMessage(
            subject='Клуб Первых',
            body='Оплатить счёт возможно в течение 5 рабочих дней. До встречи в Клубе Первых!',
            from_email=EMAIL_HOST_USER,
            to=[str(email)],
        )
        email_send.attach_file(path_payment)
        email_send.attach_file(path_pdf)
        email_send.send()
        return img_base


class FillingQuestionnaireMixin:
    """Миксин для формы анкета"""

    form = None
    contract = None
    template_name = None

    def get(self, request, contract_number):
        return render(request, self.template_name, {'form': self.form, 'contract': self.contract})

    def post(self, request, contract_number):
        form = self.form(request.POST)
        if 'docx' in request.POST:
            docx_base = form_questionnaire(self, request, contract_number)
            return render(request, 'filling_questionnaire.html', {'docx_base': docx_base, 'form': form})

        elif 'qr_code' in request.POST:
            img_base = finally_rich(self, request, contract_number)
            get_sign_img(self, request, contract_number)
            return render(request, 'filling_questionnaire.html', {'img_base': img_base})

        return render(request, 'filling_questionnaire.html', {'form': form})

# class FillingQuestionnaireMixin:
#     """Миксин для формы анкета"""
#
#     form = None
#     contract = None
#     template_name = None
#
#     def get(self, request, contract_number):
#         return render(request, self.template_name, {'form': self.form, 'contract': self.contract})
#
#     def post(self, request, contract_number):
#         global purpose, sum
#         form = self.form(request.POST)
#         contract = self.contract()
#         if form.is_valid:
#             docx = DocxTemplate(os.path.join(BASE_DIR, contract.contract_template.template_of_contract.path))
#             last_name = request.POST['last_name']
#             name = request.POST['name']
#             sur_name = request.POST['sur_name']
#             full_name = last_name.title() + ' ' + name.title() + ' ' + sur_name.title()
#             generated_date = format_date(contract.date_created, 'd MMMM yyyy', locale='ru')
#             phone = request.POST['phone']
#             passport = request.POST['passport']
#             email = request.POST['email']
#             print(get_data_from_forms(self, request, contract_number).get('last_name'))
#             print(get_data_from_forms(self, request, contract_number).get('name'))
#             if sur_name:
#                 short_name = last_name + ' ' + name[0] + '.' + sur_name[0] + '.'
#             else:
#                 short_name = last_name + ' ' + name[0] + '.'
#
#             id_contract = str(contract.number)
#             path_docx = os.path.join(BASE_DIR, 'upload\\tmp\\' + id_contract + '.docx')
#             type = str(contract.contract_template.type)
#             basic_vars = {'email': email,
#                           'full_name': full_name,
#                           'id': id_contract,
#                           'short_name': short_name,
#                           'generated_date': str(generated_date),
#                           'phone': phone,
#                           'passport': passport,
#                           }
#             sum_c = contract.amount
#             text_sum = num2text(sum_c) + ' рублей'
#             renewal_vars = {'email': email,
#                             'full_name': full_name,
#                             'id': id_contract,
#                             'short_name': short_name,
#                             'generated_date': str(generated_date),
#                             'phone': phone,
#                             'passport': passport,
#                             'sum': str(sum_c),
#                             'text_sum': text_sum
#                             }
#             if type == 'Основной':
#
#                 docx.render(basic_vars)
#
#                 docx.save(path_docx)
#             else:
#
#                 docx.render(renewal_vars)
#
#                 docx.save(path_docx)
#
#             docx_open = open(path_docx, 'rb')
#             docx_read = docx_open.read()
#             docx_base = base64.b64encode(docx_read)
#             docx_open.close()
#             os.remove(path_docx)
#             if request.method == 'POST' and 'qr_code' in request.POST:
#                 company = contract.contract_template.company
#                 generated_date_qr = format_date(contract.date_created, 'd.MM.yy', locale='ru')
#                 qr = qrcode.QRCode(
#                     version=4,
#                     error_correction=qrcode.constants.ERROR_CORRECT_L,
#                     box_size=2,
#                     border=1,
#                     )
#                 if type == 'Основной':
#                     purpose = 'Оплата участия в серии мероприятий для бизнеса "Клуба Первых". ' \
#                               'По Договору оказания услуг ФЛМ-' + id_contract + ' от ' + generated_date_qr
#                     sum = 1
#                 elif type == 'Продление':
#                     purpose = 'Оплата участия в серии мероприятий для бизнеса "Клуба Первых" ' \
#                               'Тариф "Продление" по Договору оказания услуг ПМ-'+ id_contract + ' от ' + generated_date_qr
#                     sum = contract.amount
#
#                 new_path_docx = os.path.join(BASE_DIR, 'upload\\signed_contract\\' + id_contract + '.docx')
#                 path_pdf = os.path.join(BASE_DIR, 'upload\\signed_contract\\' + id_contract + '.pdf')
#                 if company == 'ООО "КМС"':
#
#                     qr_kms = qr
#                     qr_kms.add_data(
#                         'ST00012|Name=ООО "КМС"|'
#                         'PersonalAcc=40702810838000108799|'
#                         'BankName=ПАО СБЕРБАНК|'
#                         'BIC=044525225|'
#                         'CorrespAcc=30101810400000000225|'
#                         'Purpose=' + purpose + '|'
#                         'Sum=' + str(sum) + '|'
#                         'PayeeINN=7725731508|'
#                         'KPP=772501001|'
#                         'LastName=' + last_name + '|'
#                         'FirstName=' + name + '|'
#                         'MiddleName=' + sur_name + ''
#                     )
#
#                     qr_kms.make(fit=True)
#
#                     img = qr_kms.make_image(fill_color="black", back_color="white")
#                     bytes_io = BytesIO()
#                     img.save(bytes_io, format='png')
#                     img_base = base64.b64encode(bytes_io.getvalue()).decode()
#
#                     workbook = load_workbook(os.path.join(BASE_DIR, 'Шаблон Счета на оплату КМС.xlsx'))
#                     sheet = workbook.active
#                     sheet['C8'] = purpose
#                     sheet['C10'] = str(sum)
#                     sheet['C21'] = purpose
#                     sheet['C23'] = str(sum)
#                     img_qr = Image(bytes_io)
#                     sheet.add_image(img_qr, 'B18')
#
#                     path_payment = os.path.join(BASE_DIR, 'upload\\payment\\счет_' + id_contract + '.xlsx')
#                     workbook.save(path_payment)
#                     if type == 'Основной':
#                         basic_vars['signature'] = InlineImage(docx, os.path.join(BASE_DIR, 'image (1).png'), width=Mm(20),
#                                                               height=Mm(10))
#                         docx.render(basic_vars)
#
#                         docx.save(new_path_docx)
#                     else:
#                         renewal_vars['signature'] = InlineImage(docx, os.path.join(BASE_DIR, 'image (1).png'), width=Mm(20),
#                                                               height=Mm(10))
#                         docx.render(renewal_vars)
#
#                         docx.save(new_path_docx)
#                     convertapi.api_secret = 'iP4B37Pw0h0xIn9Y'
#
#                     result = convertapi.convert('pdf', {'File': new_path_docx})
#
#                     result.file.save(path_pdf)
#                     os.remove(new_path_docx)
#                     contract.payment = path_payment
#                     contract.signed_contract = path_pdf
#                     contract.save()
#                     email = EmailMessage(
#                         subject='Клуб Первых',
#                         body='Оплатить счёт возможно в течение 5 рабочих дней. До встречи в Клубе Первых!',
#                         from_email=EMAIL_HOST_USER,
#                         to=[str(email)],
#                     )
#                     email.attach_file(path_payment)
#                     email.attach_file(path_pdf)
#                     # email.send()
#                     return render(request, self.template_name, context={'docx_base': docx_base, 'img_base': img_base})
#                 elif company == 'ООО "ДЕЛОВОЙ КЛУБ"':
#                     qr_dk = qr
#                     qr_dk.add_data(
#                         'ST00012|Name=ООО "ДЕЛОВОЙ КЛУБ"|'
#                         'PersonalAcc=40702810338000156966|'
#                         'BankName=ПАО СБЕРБАНК|'
#                         'BIC=044525225|'
#                         'CorrespAcc=30101810400000000225|'
#                         'Purpose=' + purpose + '|'
#                         'Sum=' + str(sum) + '|'
#                         'PayeeINN=9715357887|'
#                         'KPP=771501001|'
#                         'LastName=' + last_name + '|'
#                         'FirstName=' + name + '|'
#                         'MiddleName=' + sur_name + ''
#                     )
#
#                     qr_dk.make(fit=True)
#
#                     img = qr_dk.make_image(fill_color="black", back_color="white")
#                     bytes_io = BytesIO()
#                     img.save(bytes_io, format='png')
#                     img_base = base64.b64encode(bytes_io.getvalue()).decode()
#                     workbook = load_workbook(os.path.join(BASE_DIR, 'Шаблон Счета на оплату ДК.xlsx'))
#                     sheet = workbook.active
#                     sheet['C8'] = purpose
#                     sheet['C10'] = str(sum)
#                     sheet['C21'] = purpose
#                     sheet['C23'] = str(sum)
#                     img_qr = Image(bytes_io)
#                     sheet.add_image(img_qr, 'B18')
#
#                     path_payment = os.path.join(BASE_DIR, 'upload\\payment\\счет_' + id_contract + '.xlsx')
#                     workbook.save(path_payment)
#
#                     if type == 'Основной':
#                         basic_vars['signature'] = InlineImage(docx, os.path.join(BASE_DIR, 'image (1).png'), width=Mm(20),
#                                                               height=Mm(10))
#                         print(basic_vars)
#                         docx.render(basic_vars)
#                         docx.save(new_path_docx)
#                     else:
#                         renewal_vars['signature'] = InlineImage(docx, os.path.join(BASE_DIR, 'image (1).png'), width=Mm(20),
#                                                               height=Mm(10))
#                         docx.render(renewal_vars)
#                         docx.save(new_path_docx)
#
#                     convertapi.api_secret = 'iP4B37Pw0h0xIn9Y'
#
#                     result = convertapi.convert('pdf', {'File': new_path_docx})
#
#                     result.file.save(path_pdf)
#                     os.remove(new_path_docx)
#                     contract.payment = path_payment
#                     contract.signed_contract = path_pdf
#                     contract.save()
#                     email = EmailMessage(
#                         subject='Клуб Первых',
#                         body='Оплатить счёт возможно в течение 5 рабочих дней. До встречи в Клубе Первых!',
#                         from_email=EMAIL_HOST_USER,
#                         to=[str(email)],
#                     )
#                     email.attach_file(path_payment)
#                     email.attach_file(path_pdf)
#                     # email.send()
#                     return render(request, self.template_name, context={'docx_base': docx_base, 'img_base': img_base})
#
#         return render(request, self.template_name, {'form': form})

